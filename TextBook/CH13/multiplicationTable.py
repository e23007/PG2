#! python3
#python3 multiplicationTable.py <数字> -掛け算表の作成

import openpyxl,sys
from openpyxl.styles import Font

if len(sys.argv)<2:
    print('使い方：python3 multiplicationTable.py <数字>')
    sys.exit()

n=int(sys.argv[1])
wb=openpyxl.Workbook()
sheet=wb.active
normal_font=Font(size=12)
bold_font=Font(size=12,bold=True)


# 行見出し Fontと数字の入力
for i in range(1, n + 1):
    cell = sheet.cell(row=i + 1, column=1)
    cell.value = i
    cell.font = bold_font

# 列見出し Fontと数字の入力
for j in range(1, n + 1):
    cell = sheet.cell(row=1, column=j + 1)
    cell.value = j
    cell.font = bold_font

# 九九
for i in range(1, n + 1):
    for j in range(1, n + 1):
        cell = sheet.cell(row=i + 1, column=j + 1)
        cell.value = i * j
        cell.font = normal_font

wb.save('multiplicationtable.xlsx')
